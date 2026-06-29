from io import BytesIO

from openpyxl import load_workbook

from app.parsers.base import download_from_url


async def parse_xlsx(url: str) -> str:
    content = await download_from_url(url)
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 200:
                break
            rows.append(", ".join(str(cell) if cell is not None else "" for cell in row))
        sections.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))

    return "\n\n".join(sections)
